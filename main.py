import os
import sys
from pathlib import Path
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docxtpl import DocxTemplate
from openpyxl import load_workbook

class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('uis/main.ui', self)
        self.setWindowIcon(QtGui.QIcon('images/icon.png'))
        self.setWindowTitle('Kaspi Gold')

        self.btn_wrd.clicked.connect(self.execute_wrd)
        self.btn_xl.clicked.connect(self.execute_xl)

    def execute_wrd(self):
        document_path = Path(__file__).parent / "temper.docx"
        doc = DocxTemplate(document_path)
        context = { "FIO": self.lineEdit_5.text(),
                   "Number": self.lineEdit_6.text(),
                   "Cash": self.lineEdit_7.text(),
                   "Comment": self.lineEdit_8.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "generated.docx")
        os.system('start generated.docx')

    def execute_xl(self):
        fn = 'temper.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['B4'] = self.lineEdit_5.text()
        ws['B5'] = self.lineEdit_6.text()
        ws['B6'] = self.lineEdit_7.text()
        ws['B7'] = self.lineEdit_8.text()

        wb.save(Path(__file__).parent / "generated.xlsx")
        wb.close()
        os.system('start generated.xlsx')

def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
